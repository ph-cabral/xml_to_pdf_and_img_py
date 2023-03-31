from xml.dom import minidom
from script.AB import AB
from script.B import B
from script.base import base
import os
import openpyxl
from tqdm import tqdm
import time
import datetime
import calendar
import locale


# armo una lista con los xmls que estan dentro de las carpetas
ñ1 = sorted([f for f in os.listdir("./xmls/ñ1") if f.endswith(".xml")])
ñ2 = sorted([f for f in os.listdir("./xmls/ñ2") if f.endswith(".xml")])

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook() # es un objeto que instancia la clase openpyxl

# Set the locale to Spanish
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

# Crea una carpeta para guardar los archivos si no existe
if not os.path.exists('excels'):
    os.makedirs('excels')

# Ruta de la carpeta
ruta_carpeta = './excels/'

# Recorrer los XML
for k, v in {'ñ1': ñ1, 'ñ2' : ñ2}.items():
    for xml in tqdm( v, desc=k):
        # Agregar una pausa de tiempo para simular un proceso largo
        time.sleep(0.1)
        # selecciono el xml
        ruta_xml = os.path.join(f"./xmls/{k}", xml)
        # extraigo la utlima etiqueta "Documento"
        Documento = minidom.parse(ruta_xml).getElementsByTagName('Documento')[-1]
        # obtengo la fecha del cierre
        fecha_jornada = Documento.getElementsByTagName('FechaJornada')[0].firstChild.data.strip()
        fecha = datetime.datetime.strptime(fecha_jornada, '%y%m%d')
        # extraigo el mes
        mes_num = fecha.month
        # pasamos el mes de ingles a español
        mes_es = calendar.month_name[mes_num]
        
        # obtengo el numero id
        context, numero_z = base(Documento)
        condicion = Documento.getElementsByTagName('CierreDiarioInformacionDocumento')
        totales = Documento.getElementsByTagName('CierreDiarioTotales')
        importe_total = totales[0].getElementsByTagName('TotalFinal')[0].firstChild.data.strip()
        if len(condicion) == 4:
            d = AB(Documento)
            html = 'AB.html'
        elif len(condicion) == 3 or len(condicion) == 2:
            if importe_total == 0:
                d = {
                    'gravado_global': ['',0.00],
                    'no_gravado_global': ['',0.00],
                    'exento_global': ['',0.00],
                    'descuentos_global': ['',0.00],
                    'generado_global': ['',0.00],
                    'cancelados_global': ['',0.00],
                    'total_fiscal': ['',0.00],
                    'total_no_fiscal': ['',0.00]
                }
                html = 'O.html'
            else:
                d = B(Documento)
                html = 'B.html'
        else:
            d = {
                'gravado_global': 0.00,
                'no_gravado_global': ['',0.00],
                'exento_global': ['',0.00],
                'descuentos_global': ['',0.00],
                'generado_global': ['',0.00],
                'cancelados_global': ['',0.00],
                'total_fiscal': ['',0.00],
                'total_no_fiscal': ['',0.00]
            }
            html = 'O.html' 
        context.update(d)
        # seleccionar valores de cierre z
        # B
        cierre_b =  context.get('cierre_b', '')
        primer_b =  context.get('primer_b' , '')
        ultimo_b =  context.get('ultimo_b', '')
        # A
        cierre_a =  context.get('cierre_a','')
        primer_a =  context.get('primer_a','')
        ultimo_a =  context.get('ultimo_a','')
        # Global
        ng_total = round(float(context.get('gravado_global', 0)),2)
        iva10 = round(round(float(context.get('disciminacion_iva_global1', ['',0])[1]), 2),2)
        iva21 = round(float(context.get('disciminacion_iva_global2', ['',0])[1]),2)
        ng10 = round(iva10/.105,2)
        ng21 = round(iva21/.21,2)
        ivatotal = round(iva10 + iva21,2)
        total = round(float(context.get('total_final', 0)),2)


        # transcribir a excel
        row = [context['fecha_jornada'], 
               'cierre z',  
               cierre_b, primer_b, ultimo_b,
               cierre_a, primer_a, ultimo_a,
               context['pv'], 
               context['numero_z'], 
               'CONSUMIDOR FINAL',
               ng_total,
               ivatotal,
               ng10,
               iva10,
               ng21,
               iva21,
               '', '',
               total
            ] 
        
        # Comprobar si el archivo existe
        archivo_xlsx = f'./excels/{k}_{mes_es}.xlsx'
        if not os.path.exists(archivo_xlsx):
            # Si el archivo no existe, lo grabamos
            wb.save(archivo_xlsx)
            # creo un nuevo archivo
            wb = openpyxl.Workbook()
            
        # Seleccionar la hoja activa
        sheet = wb.active

        # Escribir los encabezados en la primera fila
        header = ['FECHA', 'TIPO', 'B', 'DESDE', 'HASTA', 'A', 'DESDE', 'HASTA', 'PV', 'NUMERO', 
                  'DENOMINACION', 'NG TOTAL', 'IVA TOTAL', 'NG 10.5%', 'IVA 10.5%',
                  'Nimage.pngG 21%', 'IVA 21%', 'NG 27%', 'IVA 27%', 'TOTAL']

        # Escribir los encabezados en todas las hojas
        for i, head in enumerate(header):
            sheet.cell(row=1, column=i+1, value=head)

        # escribe los valores en el archivo excel
        fecha_dia = fecha.day
        # agrego la lista a hoja
        sheet.append(row)
    
        # Guardar el archivo
        wb.save(archivo_xlsx)
       
        